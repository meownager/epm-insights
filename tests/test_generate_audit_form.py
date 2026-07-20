"""Known-answer tests for the editable Excel audit form generator."""

import io
import sys
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
from generate_audit_form import build_form, build_project_dict

RUBRIC_PATH = Path(__file__).parent.parent / "config" / "quality_rubric.yaml"
RUBRIC = yaml.safe_load(open(RUBRIC_PATH))


def _financials_row(**overrides):
    row = {
        "project_id": "90001-F",
        "billing_type": "fixed_fee",
        "estimate_total": 100000,
        "revenue_invoiced": 100000,
        "planned_margin_pct": 0.30,
        "actual_margin_pct": 0.25,
        "margin_reliability": "full",
        "internal_labor_cost": 40000,
        "external_cost_bills": 20000,
        "external_cost_expenses": 5000,
        "co_count": 1,
        "co_dollars": 5000,
        "unbilled_wip": 0,
        "overdue_invoices": 0,
        "hours_source": "logged",
    }
    row.update(overrides)
    return row


# ── build_project_dict ───────────────────────────────────────────────────────

def test_project_dict_computes_cost_from_internal_and_external():
    project = build_project_dict(_financials_row())
    assert project["cost"] == 65000  # 40000 + 20000 + 5000


def test_project_dict_revenue_matches_invoiced():
    project = build_project_dict(_financials_row())
    assert project["revenue"] == 100000


def test_project_dict_overlays_metrics_fields_when_present():
    metrics_row = {"project_name": "Widget Line Upgrade", "client": "Acme Corp",
                   "project_manager": "J. Smith"}
    project = build_project_dict(_financials_row(), metrics_row)
    assert project["project_name"] == "Widget Line Upgrade"
    assert project["client"] == "Acme Corp"
    assert project["project_manager"] == "J. Smith"


def test_project_dict_ignores_blank_metrics_fields():
    metrics_row = {"project_name": "", "client": float("nan"), "project_manager": "nan"}
    project = build_project_dict(_financials_row(), metrics_row)
    assert "project_name" not in project or project.get("project_name") in (None, "")


def test_project_dict_handles_missing_internal_labor_cost():
    project = build_project_dict(_financials_row(internal_labor_cost=""))
    assert project["cost"] == 25000  # 0 internal + 20000 + 5000


# ── build_form ───────────────────────────────────────────────────────────────

def test_build_form_writes_to_bytesio():
    project = build_project_dict(_financials_row())
    buf = io.BytesIO()
    build_form(project, RUBRIC, buf)
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb.active.title == "Audit Form"


def test_build_form_writes_to_path(tmp_path):
    project = build_project_dict(_financials_row())
    out = tmp_path / "form.xlsx"
    build_form(project, RUBRIC, out)
    assert out.exists()


def test_build_form_header_block_values():
    project = build_project_dict(_financials_row(), {"project_name": "Test Project", "client": "Test Client"})
    buf = io.BytesIO()
    build_form(project, RUBRIC, buf)
    buf.seek(0)
    ws = load_workbook(buf).active
    cells = {ws.cell(row=r, column=1).value: r for r in range(1, 15)}
    assert ws.cell(row=cells["Customer Name"], column=2).value == "Test Client"
    assert ws.cell(row=cells["Project Number"], column=2).value == "90001-F"
    assert ws.cell(row=cells["Project Name"], column=2).value == "Test Project"
    assert ws.cell(row=cells["Cost"], column=2).value == 65000
    assert ws.cell(row=cells["Revenue"], column=2).value == 100000


def test_build_form_no_celebration_category():
    project = build_project_dict(_financials_row())
    buf = io.BytesIO()
    build_form(project, RUBRIC, buf)
    buf.seek(0)
    ws = load_workbook(buf).active
    all_values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
    assert "Celebration" not in all_values


def test_build_form_score_cells_left_blank_for_epm():
    project = build_project_dict(_financials_row())
    buf = io.BytesIO()
    build_form(project, RUBRIC, buf)
    buf.seek(0)
    ws = load_workbook(buf).active
    # first rubric category row (16), score column B must be empty for the EPM to fill
    assert ws.cell(row=16, column=2).value is None


def test_build_form_autofilled_financials_present():
    project = build_project_dict(_financials_row(co_count=3, co_dollars=15000))
    buf = io.BytesIO()
    build_form(project, RUBRIC, buf)
    buf.seek(0)
    ws = load_workbook(buf).active
    cells = {ws.cell(row=r, column=1).value: r for r in range(30, 50) if ws.cell(row=r, column=1).value}
    assert ws.cell(row=cells["Change Orders"], column=2).value == 3
    assert ws.cell(row=cells["CO Dollars"], column=2).value == 15000
