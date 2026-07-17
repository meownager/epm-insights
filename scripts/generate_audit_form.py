"""Generate pre-filled, editable Excel audit forms — one per project.

Reproduces the company Project Audit Form layout: header block on top
(customer, project, PM, cost, revenue, profit %), then the category / score /
comments / rubric table. Financial fields are pre-filled from converter output;
score and comment cells are left blank for the EPM, with live Excel formulas
for Profit %, Total Score, Total Possible, and Percent.

Usage:
    python scripts/generate_audit_form.py \
        --financials data/real/audit_inputs/project_financials.csv \
        --output-dir data/real/audit_forms
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

DEFAULT_RUBRIC_PATH = Path(__file__).parent.parent / "config" / "quality_rubric.yaml"

# Layout and styling replicate the audit owner's approved form exactly
# (reference: audit_form_edited.xlsx, 2026-07-16). Do not restyle without
# a new approved reference form.
HEADER_FILL = PatternFill("solid", fgColor="1A1F2E")
TABLE_HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
LABEL_FONT = Font(bold=True, size=12)
VALUE_FONT = Font(size=12)
CATEGORY_FONT = Font(bold=True, size=12)
RUBRIC_FONT = Font(size=12, color="FF666666")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {"A": 33.9, "B": 23.5, "C": 13.1, "D": 98.3}
CATEGORY_ROW_HEIGHT = 62.0

HEADER_BLOCK_START = 2   # row 1 is the header-block table's own header row
TABLE_HEADER_ROW = 15
FIRST_CAT_ROW = 16

# All three blocks are real Excel Tables in the approved form:
# TableStyleLight8 with row AND column stripes (banded)
TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight8",
    showRowStripes=True,
    showColumnStripes=True,
    showFirstColumn=False,
    showLastColumn=False,
)


def _add_table(ws, display_name: str, ref: str, column_names: list[str]) -> None:
    cols = [TableColumn(id=i + 1, name=n) for i, n in enumerate(column_names)]
    tbl = Table(displayName=display_name, ref=ref, tableColumns=cols)
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)


def build_form(project: dict, rubric: dict, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Form"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # ── header-block table header (row 1) ──
    hc = ws.cell(row=1, column=1, value="Project Information")
    hc.font = LABEL_FONT
    vc1 = ws.cell(row=1, column=2, value="Value")
    vc1.font = VALUE_FONT

    # ── header block (rows 2-9) ──
    header_rows = [
        ("Customer Name", project.get("client", "")),
        ("Project Number", project.get("project_id", "")),
        ("Project Name", project.get("project_name", "")),
        ("Project Manager", project.get("project_manager", "")),
        ("Technical Lead", ""),
        ("Cost", project.get("cost", "")),
        ("Revenue", project.get("revenue", "")),
    ]
    r = HEADER_BLOCK_START
    for label, value in header_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        vc = ws.cell(row=r, column=2, value=value if value != "" else None)
        vc.font = VALUE_FONT
        r += 1

    cost_row, revenue_row = HEADER_BLOCK_START + 5, HEADER_BLOCK_START + 6
    pc = ws.cell(row=r, column=1, value="Profit %")
    pc.font = LABEL_FONT
    cell = ws.cell(row=r, column=2, value=f"=IF(B{revenue_row}=0,\"\",(B{revenue_row}-B{cost_row})/B{revenue_row})")
    cell.number_format = "0.0%"
    cell.font = VALUE_FONT
    r += 1

    n_cats = len(rubric["categories"])
    last_cat_row = FIRST_CAT_ROW + n_cats - 1

    ws.cell(row=r, column=1, value="Total Score").font = LABEL_FONT
    tc = ws.cell(row=r, column=2, value=f"=SUM(B{FIRST_CAT_ROW}:B{last_cat_row})")
    tc.font = VALUE_FONT
    total_score_row = r
    r += 1
    ws.cell(row=r, column=1, value="Total Possible").font = LABEL_FONT
    tp = ws.cell(row=r, column=2, value=f"=COUNT(B{FIRST_CAT_ROW}:B{last_cat_row})*4")
    tp.font = VALUE_FONT
    total_possible_row = r
    r += 1
    ws.cell(row=r, column=1, value="Percent").font = LABEL_FONT
    pct = ws.cell(
        row=r, column=2,
        value=f"=IF(B{total_possible_row}=0,\"\",B{total_score_row}/B{total_possible_row})",
    )
    pct.number_format = "0.0%"
    pct.font = VALUE_FONT

    # ── table header (row 15) ──
    for col, title in enumerate(["Category", "Score", "Comments", "Rubric"], start=1):
        c = ws.cell(row=TABLE_HEADER_ROW, column=col, value=title)
        c.fill = HEADER_FILL
        c.font = TABLE_HEADER_FONT
        c.border = BORDER

    # ── score validation: 1-4 or na ──
    dv = DataValidation(type="list", formula1='"1,2,3,4,na"', allow_blank=True)
    dv.error = "Score must be 1, 2, 3, 4, or na"
    ws.add_data_validation(dv)

    r = FIRST_CAT_ROW
    for cat in rubric["categories"]:
        anchors = "\n".join(f"{k} - {v}" for k, v in sorted(cat["anchors"].items()))
        nc = ws.cell(row=r, column=1, value=cat["name"])
        nc.font = CATEGORY_FONT
        nc.border = BORDER
        score_cell = ws.cell(row=r, column=2)
        score_cell.font = VALUE_FONT
        score_cell.border = BORDER
        dv.add(score_cell)
        cc = ws.cell(row=r, column=3)
        cc.font = VALUE_FONT
        cc.border = BORDER
        rc = ws.cell(row=r, column=4, value=anchors)
        rc.border = BORDER
        rc.alignment = Alignment(wrap_text=True, vertical="top")
        rc.font = RUBRIC_FONT
        ws.row_dimensions[r].height = CATEGORY_ROW_HEIGHT
        r += 1

    # ── pre-filled quantitative summary (starts 2 rows below the table) ──
    r = last_cat_row + 3
    autofill_header_row = r
    ws.cell(row=r, column=1, value="AUTO-FILLED FROM AUDIT DATA").font = Font(bold=True, size=12)
    ws.cell(row=r, column=2, value="Value").font = VALUE_FONT
    r += 1
    for label, key, fmt in [
        ("Billing Type", "billing_type", None),
        ("Estimate Total", "estimate_total", "#,##0.00"),
        ("Revenue Invoiced", "revenue_invoiced", "#,##0.00"),
        ("Planned Margin", "planned_margin_pct", "0.0%"),
        ("Actual Margin", "actual_margin_pct", "0.0%"),
        ("Margin Reliability", "margin_reliability", None),
        ("Change Orders", "co_count", None),
        ("CO Dollars", "co_dollars", "#,##0.00"),
        ("Unbilled WIP", "unbilled_wip", "#,##0.00"),
        ("Overdue Invoices", "overdue_invoices", None),
        ("Hours Source", "hours_source", None),
    ]:
        val = project.get(key, "")
        ws.cell(row=r, column=1, value=label).font = VALUE_FONT
        cell = ws.cell(row=r, column=2, value=val if val != "" else None)
        cell.font = VALUE_FONT
        if fmt:
            cell.number_format = fmt
        r += 1

    # ── banded Excel Tables, matching the approved form ──
    _add_table(ws, "ProjectInfo", f"A1:B{HEADER_BLOCK_START + 10}",
               ["Project Information", "Value"])
    _add_table(ws, "RubricTable", f"A{TABLE_HEADER_ROW}:D{last_cat_row}",
               ["Category", "Score", "Comments", "Rubric"])
    _add_table(ws, "AutoFilled", f"A{autofill_header_row}:B{r - 1}",
               ["AUTO-FILLED FROM AUDIT DATA", "Value"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate editable Excel audit forms")
    ap.add_argument("--financials", required=True, help="project_financials.csv from the converter")
    ap.add_argument("--output-dir", required=True, help="Folder for generated .xlsx forms")
    ap.add_argument("--rubric", default=str(DEFAULT_RUBRIC_PATH))
    args = ap.parse_args()

    rubric = yaml.safe_load(open(args.rubric))
    fin = pd.read_csv(args.financials)
    out_dir = Path(args.output_dir)

    for _, row in fin.iterrows():
        project = row.to_dict()
        # cost = planned internal+external actuals where known
        ext = pd.to_numeric(row.get("external_cost_bills"), errors="coerce") or 0
        exp = pd.to_numeric(row.get("external_cost_expenses"), errors="coerce") or 0
        internal = pd.to_numeric(row.get("internal_labor_cost"), errors="coerce")
        project["cost"] = round((internal if pd.notna(internal) else 0) + ext + exp, 2)
        project["revenue"] = row.get("revenue_invoiced", "")
        out = out_dir / f"audit_form_{row['project_id']}.xlsx"
        build_form(project, rubric, out)
        print(f"wrote {out}")

    print(f"\n{len(fin)} audit forms generated in {out_dir}")
    print("Rubric version:", rubric.get("version"))


if __name__ == "__main__":
    main()
